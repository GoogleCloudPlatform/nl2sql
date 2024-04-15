# Copyright 2024 Google LLC
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
Allows creating custom Datasets on a local, temp PGSQL/ MySQL instance
"""

import os
import typing
from functools import lru_cache

import numpy as np
import openpyxl
import pandas as pd
from google.api_core.exceptions import NotFound
from google.cloud import bigquery
from google.cloud.bigquery import SchemaField
from loguru import logger

from nl2sql.datasets.base import Dataset


@np.vectorize
def generate_pk_query(dataset_id: str, tablename: str, primary_key_column: str) -> str:
    """
    Generate DDL queries to add associated primary key columns to respective
    tables.

    Args:
        dataset_id (str): Bigquery dataset id.
        tablename (str): Bigquery table name.
        primary_key_column (str): Name of the primary key column in table.

    Returns:
        query (str): DDL query to add primary key to table.
    """
    query = (
        f"ALTER TABLE `{dataset_id}.{tablename}` "
        f"ADD PRIMARY KEY({primary_key_column}) NOT ENFORCED;"
    )
    return query


@np.vectorize
def generate_fk_query(
    dataset_id: str, tablename: str, foreign_key_column: str, references: str
) -> str:
    """
    Generate DDL queries to add associated foreign key columns to respective
    tables and their references.

    Args:
        dataset_id (str): Bigquery dataset id.
        tablename (str): Bigquery table name.
        foreign_key_column (str): Name of the foreign key column in table.
        references (str): Reference column for foreign key.

    Returns:
        query (str): DDL query to add foreign key to table.
    """
    query = (
        f"ALTER TABLE `{dataset_id}.{tablename}` "
        f"ADD FOREIGN KEY({foreign_key_column}) "
        f"REFERENCES `{dataset_id}`.{references.replace(' ','')} "
        "NOT ENFORCED;"
    )
    return query


class CustomDataset:
    """
    Instantiates a local database
    """

    def __init__(self, filepath: str, project_id: str | None, dataset_name: str):
        """
        Custom Dataset

        Args:
            filepath (str): Filepath for excel file having table schema & data.
            project_id (str | None): GCP Project Id.
            dataset_name (str): Bigquery dataset name to be created.
        """
        self.filepath = filepath
        self.dataset_name = dataset_name
        self.client = bigquery.Client(project=project_id, location="US")

    @classmethod
    @lru_cache
    def from_excel(
        cls,
        filepath: str,
        project_id: str | None = None,
        dataset_name: str = "custom_dataset",
    ) -> Dataset:
        """
        Creates and returns a Dataset object based on the specified excel file.

        Args:
            filepath (str): File path where the input excel file is located.
            project_id (str | None, optional):
                GCP Project ID where bigquery dataset is created.
                Defautls to Environment variable "GOOGLE_CLOUD_PROJECT".
            dataset_name (str, optional):
                Name of the Custom Bigquery Dataset.
                Defaults to "custom_dataset".

        Returns:
            Dataset: A Dataset Object.
        """
        if project_id is None:
            project_id = os.getenv("GOOGLE_CLOUD_PROJECT")

        custom = cls(
            filepath=filepath, project_id=project_id, dataset_name=dataset_name
        )

        dataset_id = f"{project_id}.{dataset_name}"
        try:
            dataset = custom.client.get_dataset(dataset_id)
            logger.info(f"Dataset {dataset_id} already exists.")
        except NotFound:
            dataset = bigquery.Dataset(dataset_id)
            dataset = custom.client.create_dataset(dataset)
            logger.success(f"Created dataset {dataset_id}.")

        custom.create_tables()
        custom.update_key_columns(dataset_id=dataset_id)

        return Dataset.from_connection_strings(
            name_connstr_map={
                f"{dataset_name}": f"bigquery://{project_id}/{dataset_name}"
            }
        )

    def generate_bigquery_schema(
        self, table_df: pd.DataFrame
    ) -> typing.List[SchemaField]:
        """Generate a Bigquery compatible schema from Pandas Dataframe.

        Args:
            table_df (pd.DataFrame): Table Dataframe.

        Returns:
            typing.List[SchemaField]: Bigquery Compatibel Schema.
        """
        type_mapping = {
            "i": "INTEGER",
            "u": "NUMERIC",
            "b": "BOOLEAN",
            "f": "FLOAT",
            "O": "STRING",
            "S": "STRING",
            "U": "STRING",
            "M": "TIMESTAMP",
        }
        schema = []
        for column, dtype in table_df.dtypes.items():
            val = table_df[column].iloc[0]
            mode = "REPEATED" if isinstance(val, list) else "NULLABLE"

            if isinstance(val, dict) or (
                mode == "REPEATED" and isinstance(val[0], dict)
            ):
                fields = self.generate_bigquery_schema(pd.json_normalize(val))
            else:
                fields = []

            type_ = "RECORD" if fields else type_mapping.get(dtype.kind)
            schema.append(
                SchemaField(
                    name=column,  # type: ignore
                    field_type=type_,  # type: ignore
                    mode=mode,
                    fields=fields,
                )
            )
        return schema

    def create_tables(self):
        """
        Create Tables in Bigquery based on the sheetname in excel file.
        """
        workbook = openpyxl.load_workbook(self.filepath)
        sheetnames = workbook.sheetnames
        sheetnames = [
            sheetname
            for sheetname in sheetnames
            if sheetname not in ["Primary Keys", "Foreign Keys"]
        ]
        for sheetname in sheetnames:
            table_id = f"{self.dataset_name}.{sheetname}"
            table_df = pd.read_excel(self.filepath, sheet_name=sheetname)
            table_df = table_df.convert_dtypes()
            schema = self.generate_bigquery_schema(table_df)
            job_config = bigquery.LoadJobConfig(
                schema=schema, write_disposition="WRITE_TRUNCATE"
            )
            job = self.client.load_table_from_dataframe(
                table_df, table_id, job_config=job_config
            )
            job.result()
            logger.success(f"Created table {table_id}")

    def update_key_columns(self, dataset_id):
        """
        Update Key columns of tables present in the dataset.
        """
        try:
            pkey = pd.read_excel(self.filepath, sheet_name="Primary Keys")
            fkey = pd.read_excel(self.filepath, sheet_name="Foreign Keys")

            pkey["Query"] = generate_pk_query(
                dataset_id, pkey["Table"], pkey["Primary Key"]
            )
            fkey["Query"] = generate_fk_query(
                dataset_id, fkey["Table"], fkey["Foreign Key"], fkey["References"]
            )

            for pk_query in pkey["Query"].tolist():
                self.client.query(pk_query)

            for fk_query in fkey["Query"].tolist():
                self.client.query(fk_query)
        except ValueError as err:
            logger.error(f"Sheetname value error: {err}")
